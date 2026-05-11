from openpyxl import load_workbook
from app.utils.logger import setup_logger
logger = setup_logger(__name__)


def load_excel(file_path: str):
    logger.info(f"Loading Excel File: {file_path}")
    workbook = load_workbook(
        filename=file_path,
        data_only=True
    )
    logger.info(
        f"Workbook Loaded Successfully: {workbook.sheetnames}"
    )
    return workbook